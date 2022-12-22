
import json
import openpyxl



class Student:
    def __init__(self):
        self.x_file='student_database/database/students_data.json'
        try :
            with open(self.x_file) as s_obj:
                self.x_data=json.load(s_obj)
                s_obj.close()
        except json.JSONDecodeError:
            self.__c_student_tab()
            
    def __c_student_tab(self):
        with open(self.x_file,'w') as s_obj:
            temp={
                "students":[

                ]
            }
            json.dump(temp,s_obj,indent=2)
            s_obj.close()
        self.__init__()


    def __show(self):
        for key in self.x_data['students']:
            print(key)
    def __update(self,list,student):
        student=student.title()
        self.x_state=True
        index=self.__show_details(student)
        try:
            for i in list:
                temp=self.x_data['students'][index][student]
                temp[i]=input(f'Enter {student} {i} :')
            self.x_data['students'][index][student].update(temp)
            self.__add(self.x_data,student)
        except TypeError:
            print(f'{student} is not databse')


    def update(self,list):
        """
        update student details by entering student name &
        insert whta you want to update
        """
        self.x_state=False
        st_name=input('Enter student name to update : ')
        self.__show_details(st_name)
        self.__update(list,st_name)
    def delete_student(self):
        """
        Entering students name for deleting"""
        self.__show()
        student=input('Enter student name : ')
        student=student.title()
        len_data=len(self.x_data['students'])
        for i in range(len_data):
            for key,value in self.x_data['students'][i-1].items():
        
                if key==student:
                    del self.x_data['students'][i-1]            
        len_data+=1
        with open(self.x_file,'w') as s_obj:
            json.dump(self.x_data,s_obj,indent=2)
            print(f"successfully deleted {str(student)}")
            s_obj.close()

    def __show_details(self,student):
        student=student.title()
        lent=len(self.x_data['students'])
        try:

            for i in range(lent):
                for key,value in self.x_data['students'][i-1].items():
                    if key==student:
                        if self.x_state==True:
                            return int(i-1)
                        else:
                            print(student)
                            for keys,values in self.x_data['students'][i-1][f'{student}'].items():
                                print(f'\t{keys} : {values}')
                    else:
                        print(f'{student} not in databse...check name or add')
        except AttributeError:
            print(student)
            for keys,values in self.x_data['students'][i-1][f'{student}'].items():
                print(f'\t{keys} : {values}')
    def details(self):
        """
        Enter by enteriing student name 
        to get student details"""
        st=input('Enter student name :')
        self.__show_details(st)

    def __add(self,data,st_name):
        with open(self.x_file,'w') as s_obj:
            json.dump(data,s_obj,indent=2)
            try:

                if self.x_state:
                    print(f"successfully updated {str(st_name)}")
                else:
                    print(f"successfully added {str(st_name)}")
            except AttributeError:
                print(f"successfully added {str(st_name)}")
            s_obj.close()
    def __merge_cells(self,ws):
        self.x_alp=['A','B','C','D','E','F','G','H','I','J','K']
        self.x_lent=len([ k for k in self.x_data['students']])
        print(self.x_lent)
        ws["A1"].value='students'
        self.__st_name(ws)
        for i in range(1,self.x_lent+2):
            ws.merge_cells(f"A{i}:B{i}")
    def __title(self,file):
        i=2
        cell=2
        titles=[]
        for key,value in self.x_data['students'][0].items():
    
            for k,v in value.items():
                titles.append(k)

        for j in titles:
            file[f"{self.x_alp[cell]}1"].value=j
            i+=2
            cell+=2
        i=2
        cell=2
        for k in titles:
            file.merge_cells(f"{self.x_alp[i]}1:{self.x_alp[i+1]}1")
            i+=2
            cell+=2
    def __st_details(self,file):
        num=1
        alp=2
        alpn=2
        for i in range(self.x_lent):
            for key,values in self.x_data['students'][num-1][self.x_students[num-1]].items():
                file[f'{self.x_alp[alp]}{alpn}'].value=values
            
                alp+=2
            alp=2
            num+=1
            alpn+=1
    def __st_name(self,file):
        num=1
        self.x_students=[]
        for n in range(self.x_lent):
            for k,v in self.x_data['students'][num-1].items():
                self.x_students.append(k)
                num+=1
        num=2
        for i in self.x_students:
            file[f'A{num}']=i
            num+=1
        self.__st_details(file)
        

    def student_excel(self):
        """
        get all students list in excel form"""
        wb=openpyxl.Workbook()
        #print(wb.sheetnames)
        del wb['Sheet']
        ws=wb.create_sheet('students')

        self.__merge_cells(ws)
        self.__title(ws)

       

        wb.save('student_database\list students\student_list.xlsx')
        print('successfully created (stdents_list.xlsx) file in student_database\list students\student_list.xlsx ')

    def __get_input(self,l_data):
        self.x_name=input('Enter student name :')
        self.x_name=self.x_name.title()
        self.x_j_s_data={
            f"{self.x_name}":{

            }
        }
        for i in l_data:
            details=input(f"Enter {self.x_name} {i} :")
            if details.isalnum():
                details=details.title()
            else:
                details=details
            self.x_j_s_data[f"{self.x_name}"][f"{i}"]=details
    def __temp_data(self,data):
        try:
            with open(self.x_file,'w') as s_obj:
                temp=self.x_data['students']
                temp.append(data)
                self.__add(self.x_data,self.x_name)
                s_obj.close()
        except json.decoder.JSONDecodeError:
            self.__c_student_tab()


    def add_student(self,list) -> list :
        """
        Insert data variables\n
        in list form\n
        EX: [name,section,phone_number]
        """
        self.__get_input(list)
        '''self.x_j_s_data={
            str(self.x_name):[
                {
                    "name":self.x_name,
                    "section":self.section,
                    "dob":self.dob,
                    "phone":self.phone
                }
            ]
        
        }'''
        len_data=len(self.x_data['students'])
        for i in range(len_data):
            for key,value in self.x_data['students'][i-1].items():
                if key==str(self.x_name):
                    print(f'{key} {value} is aldready exits Try Again...')
                    return 0
            len_data+=1
        '''def __js_st_data(self,)'''
        self.__temp_data(self.x_j_s_data)

                
